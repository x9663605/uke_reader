# -*- coding: utf-8 -*-
"""
/***************************************************************************
 UkeReader
                                 A QGIS plugin
 UKE data reader
                              -------------------
        begin                : 2015-11-19
        git sha              : $Format:%H$
        copyright            : (C) 2015 by x9663605
        email                : x9663605@gmail.com
 ***************************************************************************/

/***************************************************************************
 *                                                                         *
 *   This program is free software; you can redistribute it and/or modify  *
 *   it under the terms of the GNU General Public License as published by  *
 *   the Free Software Foundation; either version 2 of the License, or     *
 *   (at your option) any later version.                                   *
 *                                                                         *
 ***************************************************************************/
"""
from PyQt4.QtCore import QSettings, QTranslator, qVersion, QCoreApplication, QVariant
from PyQt4.QtGui import QAction, QIcon, QFileDialog
# Initialize Qt resources from file resources.py
import resources
# Import the code for the dialog
from uke_reader_dialog import UkeReaderDialog
import os.path

import sys
sys.path.append(os.path.join(os.path.dirname(__file__), "openpyxl"))
sys.path.append(os.path.join(os.path.dirname(__file__), "jdcal"))
sys.path.append(os.path.join(os.path.dirname(__file__), "et_xmlfile"))

from openpyxl import *
from openpyxl.cell import get_column_letter, column_index_from_string
import re
import locale
from qgis.core import *

def get_coordinates(col_name, start_row, worksheet, accuracy):
    """
    Get, transform to decimal, round with accuracy
    """
    coordinates = []
    range_expr = "{col}{start_row}:{col}{end_row}".format(col=col_name, start_row=start_row, end_row=worksheet.max_row)

    for row in worksheet.iter_rows(range_string=range_expr):
        if len(row) > 0:
            cell = row[0]
            sliced = re.search("(\d+)[ENSW](\d+)\'(\d+)\"", cell.value)
            degrees = int(sliced.group(1))
            minutes = float(sliced.group(2)) / 60
            seconds = float(sliced.group(3)) / 3600
            coordinate = round(degrees + minutes + seconds, accuracy)
            coordinates.append(coordinate)
        else:
            continue
    return coordinates


class UkeReader:
    """QGIS Plugin Implementation."""

    def __init__(self, iface):
        """Constructor.

        :param iface: An interface instance that will be passed to this class
            which provides the hook by which you can manipulate the QGIS
            application at run time.
        :type iface: QgsInterface
        """
        # Save reference to the QGIS interface
        self.iface = iface
        # initialize plugin directory
        self.plugin_dir = os.path.dirname(__file__)
        # initialize locale
        locale = QSettings().value('locale/userLocale')[0:2]
        locale_path = os.path.join(
            self.plugin_dir,
            'i18n',
            'UkeReader_{}.qm'.format(locale))

        if os.path.exists(locale_path):
            self.translator = QTranslator()
            self.translator.load(locale_path)

            if qVersion() > '4.3.3':
                QCoreApplication.installTranslator(self.translator)

        # Create the dialog (after translation) and keep reference
        self.dlg = UkeReaderDialog()

        # Declare instance attributes
        self.actions = []
        self.menu = self.tr(u'&UKE Reader')
        # TODO: We are going to let the user set this up in a future iteration
        self.toolbar = self.iface.addToolBar(u'UkeReader')
        self.toolbar.setObjectName(u'UkeReader')

        self.dlg.lineEdit_2.setText("E")
        self.dlg.lineEdit_3.setText("F")
        self.dlg.spinBox.setValue(2)
        self.dlg.pushButton.clicked.connect(self.select_input_file)

    # noinspection PyMethodMayBeStatic
    def tr(self, message):
        """Get the translation for a string using Qt translation API.

        We implement this ourselves since we do not inherit QObject.

        :param message: String for translation.
        :type message: str, QString

        :returns: Translated version of message.
        :rtype: QString
        """
        # noinspection PyTypeChecker,PyArgumentList,PyCallByClass
        return QCoreApplication.translate('UkeReader', message)


    def add_action(
        self,
        icon_path,
        text,
        callback,
        enabled_flag=True,
        add_to_menu=True,
        add_to_toolbar=True,
        status_tip=None,
        whats_this=None,
        parent=None):
        """Add a toolbar icon to the toolbar.

        :param icon_path: Path to the icon for this action. Can be a resource
            path (e.g. ':/plugins/foo/bar.png') or a normal file system path.
        :type icon_path: str

        :param text: Text that should be shown in menu items for this action.
        :type text: str

        :param callback: Function to be called when the action is triggered.
        :type callback: function

        :param enabled_flag: A flag indicating if the action should be enabled
            by default. Defaults to True.
        :type enabled_flag: bool

        :param add_to_menu: Flag indicating whether the action should also
            be added to the menu. Defaults to True.
        :type add_to_menu: bool

        :param add_to_toolbar: Flag indicating whether the action should also
            be added to the toolbar. Defaults to True.
        :type add_to_toolbar: bool

        :param status_tip: Optional text to show in a popup when mouse pointer
            hovers over the action.
        :type status_tip: str

        :param parent: Parent widget for the new action. Defaults None.
        :type parent: QWidget

        :param whats_this: Optional text to show in the status bar when the
            mouse pointer hovers over the action.

        :returns: The action that was created. Note that the action is also
            added to self.actions list.
        :rtype: QAction
        """

        icon = QIcon(icon_path)
        action = QAction(icon, text, parent)
        action.triggered.connect(callback)
        action.setEnabled(enabled_flag)

        if status_tip is not None:
            action.setStatusTip(status_tip)

        if whats_this is not None:
            action.setWhatsThis(whats_this)

        if add_to_toolbar:
            self.toolbar.addAction(action)

        if add_to_menu:
            self.iface.addPluginToMenu(
                self.menu,
                action)

        self.actions.append(action)

        return action

    def initGui(self):
        """Create the menu entries and toolbar icons inside the QGIS GUI."""

        icon_path = ':/plugins/UkeReader/icon.png'
        self.add_action(
            icon_path,
            text=self.tr(u'UKE Reader'),
            callback=self.run,
            parent=self.iface.mainWindow())


    def unload(self):
        """Removes the plugin menu item and icon from QGIS GUI."""
        for action in self.actions:
            self.iface.removePluginMenu(
                self.tr(u'&UKE Reader'),
                action)
            self.iface.removeToolBarIcon(action)
        # remove the toolbar
        del self.toolbar

    def select_input_file(self):
        filename = QFileDialog.getOpenFileName(None,
                "Open XLSX file", '',
                "XLSX file (*.xlsx)")
        self.dlg.lineEdit.setText(filename)

    def run(self):
        """Run method that performs all the real work"""
        # show the dialog
        lan = locale.getdefaultlocale()
        if lan[0] == "pl_PL":
            self.dlg.label.setText(u"Arkusz:")
            self.dlg.label_2.setText(u"Kolumna X:")
            self.dlg.label_3.setText(u"Kolumna Y:")
            self.dlg.label_4.setText(u"Numer pierwszego wiersza:")
        self.dlg.show()
        # Run the dialog event loop
        result = self.dlg.exec_()
        # See if OK was pressed
        if result:
            filename = self.dlg.lineEdit.text()
            columnx = self.dlg.lineEdit_2.text()
            columny = self.dlg.lineEdit_3.text()
            startrow = int(self.dlg.spinBox.value())
            wb = load_workbook(filename)
            ws = wb.active
            x = get_coordinates(columnx, startrow, ws, 10)
            y = get_coordinates(columny, startrow, ws, 10)

            vl = QgsVectorLayer("Point?crs=EPSG:4326", "uke_coordinates", "memory")
            pr = vl.dataProvider()
            vl.startEditing()
            pr.addAttributes([QgsField("X", QVariant.Double),
                QgsField("Y", QVariant.Double)])

            columns = column_index_from_string(get_column_letter(ws.get_highest_column()))
            for i in range(1, columns+1):
                col_name = get_column_letter(i)
                if col_name not in [columnx, columny]:
                    pr.addAttributes([QgsField(unicode(ws.cell(row=startrow-1,column=i).value), QVariant.String)])

            for i in range(0, ws.max_row-1):
                fet = QgsFeature()
                fet.setGeometry(QgsGeometry.fromPoint(QgsPoint(x[i], y[i])))
                li = [x[i], y[i]]
                for j in range(1, columns+1):
                    col_name = get_column_letter(j)
                    if col_name not in [columnx, columny]:
                        li.extend([unicode(ws.cell(row=i+2,column=j).value)])
                fet.setAttributes(li)
                pr.addFeatures([fet])
                vl.updateExtents()

            QgsMapLayerRegistry.instance().addMapLayer(vl)
            vl.commitChanges()
